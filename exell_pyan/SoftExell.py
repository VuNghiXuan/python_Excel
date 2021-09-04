import xlwings as xw
import openpyxl
import pandas as pd
import os
class Workbook():
	def __init__(self, path, name):
		self.path = path
		self.name = name
		# self.sheets = sheets
		self.active = False

		

	def open(self):
		xw.Book(self.path + self.name)
		self.active = True

def main():
	# Information file_infor
	file_infor="data.txt"
	file_name = "test.xlsx"
	# Infor_workbook
	file_path = os.getcwd()#"D:\\ThanhVu\\python\\pyExell\\" #|||
	fullName_file = file_path+file_name

	print("fullName_file", fullName_file)

	wb=xw.Book(fullName_file)
	sh_mun=xw.sheets
	print ("----", sh_mun)
	# wb = Workbook(file_path, file_name)
	# app = xw.App()
	# app.books.open[fullName_file]
	# wb.open()
	count_sh = wb.api.Sheets.Count
	# df = pd.read_excel(wb, "Sheet1")
	# print (len(df))
	# wb = openpyxl.load_workbook(wb.path+) wb
	# count_sh = len(wb_open.sheetnames)
	# print ("----", count_sh)
main()
# path = "I:\\\\Python\\pyExell\\test.xlsx"
# wb = xl.Book(path)
# sht = wb.sheets[1]

# UsedRange of address
# useRange = sht.api.UsedRange.Address
# a_range = sht.used_range.address
# # Cách 1: last_row
# from xlwings import Range, constants

# lr = wb.sheets[1].range('A' + str(wb.sheets[1].cells.last_cell.row)).end('up').row
# print(lr)

# Cách 2
# print (useRangeaddress, a_range)
# wb.Sheets.Count